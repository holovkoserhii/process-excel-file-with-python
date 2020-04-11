import openpyxl as xl


def process_workbook(filename):
    workbook = xl.load_workbook(f'{filename}.xlsx')
    print(workbook)
    sheet = workbook['Sheet1']

    # cell = sheet['a1']
    # either this way!
    # cell = sheet.cell(1, 1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # saving in a different file
    workbook.save(f'{filename}_result.xlsx')
